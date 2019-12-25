'''The MIT License (MIT)
Copyright (c) Microsoft
Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
copies of the Software, and to permit persons to whom the Software is
furnished to do so, subject to the following conditions:
The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.
THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.'''
from pykinect2 import PyKinectV2
from pykinect2.PyKinectV2 import *

import ctypes
import _ctypes
import pygame
import sys

import numpy
if sys.hexversion >= 0x03000000:
    import _thread as thread
else:
    import thread


from pygame.locals import *

from pgu import gui
import time, sched
import threading
from socket import *
from  datetime  import  datetime

import cv2 

import heapq
from collections import namedtuple
from time import monotonic as _time

import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook

KINECT_MAX_BODY_COUNT = 6

class PyKinectRuntime(object):
    """manages Kinect objects and simplifying access to them"""
    def __init__(self, frame_source_types):
        # recipe to get address of surface: http://archives.seul.org/pygame/users/Apr-2008/msg00218.html
        is_64bits = sys.maxsize > 2**32
        if not is_64bits:
           self.Py_ssize_t = ctypes.c_int
        else:
           self.Py_ssize_t = ctypes.c_int64

        self._PyObject_AsWriteBuffer = ctypes.pythonapi.PyObject_AsWriteBuffer
        self._PyObject_AsWriteBuffer.restype = ctypes.c_int
        self._PyObject_AsWriteBuffer.argtypes = [ctypes.py_object,
                                          ctypes.POINTER(ctypes.c_void_p),
                                          ctypes.POINTER(self.Py_ssize_t)]
        
        #self._color_frame_ready = PyKinectV2._event()
        #self._depth_frame_ready = PyKinectV2._event()
        #self._body_frame_ready = PyKinectV2._event()
        #self._body_index_frame_ready = PyKinectV2._event()
        #self._infrared_frame_ready = PyKinectV2._event()
        #self._long_exposure_infrared_frame_ready = PyKinectV2._event()
        #self._audio_frame_ready = PyKinectV2._event()

        self._close_event = ctypes.windll.kernel32.CreateEventW(None, False, False, None)

        self._color_frame_arrived_event = 0
        self._depth_frame_arrived_event = 0
        self._body_frame_arrived_event = 0
        self._body_index_frame_arrived_event = 0
        self._infrared_frame_arrived_event = 0  
        self._long_exposure_infrared_frame_arrived_event = 0
        self._audio_frame_arrived_event = 0

        self._color_frame_lock = thread.allocate()
        self._depth_frame_lock = thread.allocate()
        self._body_frame_lock = thread.allocate()
        self._body_index_frame_lock = thread.allocate()
        self._infrared_frame_lock = thread.allocate()
        self._long_exposure_infrared_frame_lock = thread.allocate()
        self._audio_frame_lock = thread.allocate()

        #initialize sensor
        self._sensor = ctypes.POINTER(PyKinectV2.IKinectSensor)()
        hres = ctypes.windll.kinect20.GetDefaultKinectSensor(ctypes.byref(self._sensor)) 
        hres = self._sensor.Open() 

        self._mapper = self._sensor.CoordinateMapper

        self.frame_source_types = frame_source_types
        self.max_body_count = KINECT_MAX_BODY_COUNT

        self._handles = (ctypes.c_voidp * 8)()
        self._handles[0] = self._close_event
        self._handles[1] = self._close_event
        self._handles[2] = self._close_event
        self._handles[3] = self._close_event
        self._handles[4] = self._close_event
        self._handles[5] = self._close_event
        self._handles[6] = self._close_event
        self._handles[7] = self._close_event

        self._waitHandleCount = 1

        self._color_source = self._sensor.ColorFrameSource 
        self.color_frame_desc = self._color_source.FrameDescription
        self._infrared_source = self._sensor.InfraredFrameSource
        self.infrared_frame_desc = self._infrared_source.FrameDescription 
        self._depth_source = self._sensor.DepthFrameSource 
        self.depth_frame_desc = self._depth_source.FrameDescription 
        self._body_index_source = self._sensor.BodyIndexFrameSource 
        self.body_index_frame_desc = self._body_index_source.FrameDescription 
        self._body_source = self._sensor.BodyFrameSource 
        self._body_frame_data = ctypes.POINTER(ctypes.POINTER(IBody))
        self.max_body_count = self._body_source.BodyCount

        self._color_frame_data = None 
        self._depth_frame_data = None 
        self._body_frame_data = None
        self._body_index_frame_data = None
        self._infrared_frame_data = None
        self._long_exposure_infrared_frame_data = None
        self._audio_frame_data = None

        if(self.frame_source_types & FrameSourceTypes_Color):
            self._color_frame_data = ctypes.POINTER(ctypes.c_ubyte) 
            self._color_frame_data_capacity = ctypes.c_uint(self.color_frame_desc.Width * self.color_frame_desc.Height * 4)
            self._color_frame_data_type = ctypes.c_ubyte * self._color_frame_data_capacity.value
            self._color_frame_data = ctypes.cast(self._color_frame_data_type(), ctypes.POINTER(ctypes.c_ubyte))
            self._color_frame_reader = self._color_source.OpenReader()
            self._color_frame_arrived_event = self._color_frame_reader.SubscribeFrameArrived()
            self._handles[self._waitHandleCount] = self._color_frame_arrived_event
            self._waitHandleCount += 1

        if(self.frame_source_types & FrameSourceTypes_Infrared):
            self._infrared_frame_data = ctypes.POINTER(ctypes.c_ushort) 
            self._infrared_frame_data_capacity = ctypes.c_uint(self.infrared_frame_desc.Width * self.infrared_frame_desc.Height)
            self._infrared_frame_data_type = ctypes.c_ushort * self._infrared_frame_data_capacity.value
            self._infrared_frame_data = ctypes.cast(self._infrared_frame_data_type(), ctypes.POINTER(ctypes.c_ushort))
            self._infrared_frame_reader = self._infrared_source.OpenReader()
            self._infrared_frame_arrived_event = self._infrared_frame_reader.SubscribeFrameArrived()
            self._handles[self._waitHandleCount] = self._infrared_frame_arrived_event
            self._waitHandleCount += 1
            
        if(self.frame_source_types & FrameSourceTypes_Depth):
            self._depth_frame_data = ctypes.POINTER(ctypes.c_ushort) 
            self._depth_frame_data_capacity = ctypes.c_uint(self.depth_frame_desc.Width * self.depth_frame_desc.Height)
            self._depth_frame_data_type = ctypes.c_ushort * self._depth_frame_data_capacity.value
            self._depth_frame_data = ctypes.cast(self._depth_frame_data_type(), ctypes.POINTER(ctypes.c_ushort))
            self._depth_frame_reader = self._depth_source.OpenReader()
            self._depth_frame_arrived_event = self._depth_frame_reader.SubscribeFrameArrived()
            self._handles[self._waitHandleCount] = self._depth_frame_arrived_event
            self._waitHandleCount += 1

        if(self.frame_source_types & FrameSourceTypes_BodyIndex):
            self._body_index_frame_data = ctypes.POINTER(ctypes.c_ubyte) 
            self._body_index_frame_data_capacity = ctypes.c_uint(self.body_index_frame_desc.Width * self.body_index_frame_desc.Height)
            self._body_index_frame_data_type = ctypes.c_ubyte * self._body_index_frame_data_capacity.value
            self._body_index_frame_data = ctypes.cast(self._body_index_frame_data_type(), ctypes.POINTER(ctypes.c_ubyte))
            self._body_index_frame_reader = self._body_index_source.OpenReader()
            self._body_index_frame_arrived_event = self._body_index_frame_reader.SubscribeFrameArrived()
            self._handles[self._waitHandleCount] = self._body_index_frame_arrived_event
            self._waitHandleCount += 1

        self._body_frame_data = None 
        if(self.frame_source_types & FrameSourceTypes_Body):
            self._body_frame_data_capacity = ctypes.c_uint(self.max_body_count)
            self._body_frame_data_type = ctypes.POINTER(IBody) * self._body_frame_data_capacity.value
            self._body_frame_data = ctypes.cast(self._body_frame_data_type(), ctypes.POINTER(ctypes.POINTER(IBody)))
            self._body_frame_reader = self._body_source.OpenReader()
            self._body_frame_arrived_event = self._body_frame_reader.SubscribeFrameArrived()
            self._body_frame_bodies = None
            self._handles[self._waitHandleCount] = self._body_frame_arrived_event
            self._waitHandleCount += 1

        thread.start_new_thread(self.kinect_frame_thread, ())

        self._last_color_frame = None
        self._last_depth_frame = None
        self._last_body_frame = None
        self._last_body_index_frame = None
        self._last_infrared_frame = None
        self._last_long_exposure_infrared_frame = None
        self._last_audio_frame = None

        start_clock = time.process_time()
        self._last_color_frame_access = self._last_color_frame_time = start_clock
        self._last_body_frame_access = self._last_body_frame_time = start_clock
        self._last_body_index_frame_access = self._last_body_index_frame_time = start_clock
        self._last_depth_frame_access = self._last_depth_frame_time = start_clock
        self._last_infrared_frame_access = self._last_infrared_frame_time = start_clock
        self._last_long_exposure_infrared_frame_access = self._last_long_exposure_infrared_frame_time = start_clock
        self._last_audio_frame_access = self._last_audio_frame_time = start_clock

    def close(self):
        if self._sensor is not None:
            ctypes.windll.kernel32.SetEvent(self._close_event)
            ctypes.windll.kernel32.CloseHandle(self._close_event)

            self._color_frame_reader = None
            self._depth_frame_reader = None
            self._body_index_frame_reader = None
            self._body_frame_reader = None

            self._color_source = None
            self._depth_source = None
            self._body_index_source = None
            self._body_source = None

            self._body_frame_data = None

            self._sensor.Close()
            self._sensor = None

    def __del__(self):
        self.close()

    def __enter__(self):
        return self

    def __exit__(self, *args):
        self.close()

    def surface_as_array(self, surface_buffer_interface):
       address = ctypes.c_void_p()
       size = self.Py_ssize_t()
       self._PyObject_AsWriteBuffer(surface_buffer_interface,
                              ctypes.byref(address), ctypes.byref(size))
       bytes = (ctypes.c_byte * size.value).from_address(address.value)
       bytes.object = surface_buffer_interface
       return bytes

    def has_new_color_frame(self):
        has = (self._last_color_frame_time > self._last_color_frame_access)
        return has

    def has_new_depth_frame(self):
        has = (self._last_depth_frame_time > self._last_depth_frame_access)
        return has

    def has_new_body_frame(self):
        has = (self._last_body_frame_time > self._last_body_frame_access)
        return has

    def has_new_body_index_frame(self):
        has = (self._last_body_index_frame_time > self._last_body_index_frame_access)
        return has

    def has_new_infrared_frame(self):
        has = (self._last_infrared_frame_time > self._last_infrared_frame_access)
        return has

    def has_new_long_exposure_infrared_frame(self):
        has = (self._last_long_exposure_infrared_frame_time > self._last_long_exposure_infrared_frame_access)
        return has

    def has_new_audio_frame(self):
        has = (self._last_audio_frame_time > self._last_audio_frame_access)
        return has


    def get_last_color_frame(self):
        with self._color_frame_lock:
            if self._color_frame_data is not None:
                data = numpy.copy(numpy.ctypeslib.as_array(self._color_frame_data, shape=(self._color_frame_data_capacity.value,)))
                self._last_color_frame_access = time.process_time()
                return data
            else:
                return None

    def get_last_infrared_frame(self):
        with self._infrared_frame_lock:
            if self._infrared_frame_data is not None:
                data = numpy.copy(numpy.ctypeslib.as_array(self._infrared_frame_data, shape=(self._infrared_frame_data_capacity.value,)))
                self._last_infrared_frame_access = time.process_time()
                return data
            else:
                return None

    def get_last_depth_frame(self):
        with self._depth_frame_lock:
            if self._depth_frame_data is not None:
                data = numpy.copy(numpy.ctypeslib.as_array(self._depth_frame_data, shape=(self._depth_frame_data_capacity.value,)))
                self._last_depth_frame_access = time.process_time()
                return data
            else:
                return None

    def get_last_body_index_frame(self):
        with self._body_index_frame_lock:
            if self._body_index_frame_data is not None:
                data = numpy.copy(numpy.ctypeslib.as_array(self._body_index_frame_data, shape=(self._body_index_frame_data_capacity.value,)))
                self._last_body_index_frame_access = time.process_time()
                return data
            else:
                return None

    def get_last_body_frame(self):
        with self._body_frame_lock:
            if self._body_frame_bodies is not None:
                self._last_body_frame_access = time.process_time()
                return self._body_frame_bodies.copy()
            else:
                return None


    def body_joint_to_color_space(self, joint): 
        return self._mapper.MapCameraPointToColorSpace(joint.Position) 

    def body_joint_to_depth_space(self, joint): 
        return self._mapper.MapCameraPointToDepthSpace(joint.Position) 


    def body_joints_to_color_space(self, joints):
        joint_points = numpy.ndarray((PyKinectV2.JointType_Count), dtype=numpy.object)

        for j in range(0, PyKinectV2.JointType_Count):
            joint_points[j] = self.body_joint_to_color_space(joints[j])

        return joint_points

    def body_joints_to_depth_space(self, joints):
        joint_points = numpy.ndarray((PyKinectV2.JointType_Count), dtype=numpy.object)

        for j in range(0, PyKinectV2.JointType_Count):
            joint_points[j] = self.body_joint_to_depth_space(joints[j])

        return joint_points

    def kinect_frame_thread(self):
        while 1:    
                wait = ctypes.windll.kernel32.WaitForMultipleObjects(self._waitHandleCount, self._handles, False, PyKinectV2._INFINITE)
               
                if wait == 0: 
                    break
                
                if self._handles[wait] == self._color_frame_arrived_event: 
                    self.handle_color_arrived(wait)
                elif self._handles[wait] == self._depth_frame_arrived_event: 
                    self.handle_depth_arrived(wait)
                elif self._handles[wait] == self._body_frame_arrived_event: 
                    self.handle_body_arrived(wait)
                elif self._handles[wait] == self._body_index_frame_arrived_event: 
                    self.handle_body_index_arrived(wait)
                elif self._handles[wait] == self._infrared_frame_arrived_event: 
                    self.handle_infrared_arrived(wait)
                elif self._handles[wait] == self._long_exposure_infrared_frame_arrived_event: 
                    self.handle_long_exposure_infrared_arrived(wait)
                elif self._handles[wait] == self._audio_frame_arrived_event: 
                    self.handle_audio_arrived(wait)
                else:
                    break

    
    def handle_color_arrived(self, handle_index):
        colorFrameEventData = self._color_frame_reader.GetFrameArrivedEventData(self._handles[handle_index])
        colorFrameRef = colorFrameEventData.FrameReference
        try:
            colorFrame = colorFrameRef.AcquireFrame()
            try:
                with self._color_frame_lock:
                    colorFrame.CopyConvertedFrameDataToArray(self._color_frame_data_capacity, self._color_frame_data, PyKinectV2.ColorImageFormat_Bgra)
                    self._last_color_frame_time = time.process_time()
            except: 
                pass
            colorFrame = None
        except:
            pass
        colorFrameRef = None
        colorFrameEventData = None


    def handle_depth_arrived(self, handle_index):
        depthFrameEventData = self._depth_frame_reader.GetFrameArrivedEventData(self._handles[handle_index])
        depthFrameRef = depthFrameEventData.FrameReference
        try:
            depthFrame = depthFrameRef.AcquireFrame()
            try:
                with self._depth_frame_lock:
                    depthFrame.CopyFrameDataToArray(self._depth_frame_data_capacity, self._depth_frame_data)
                    self._last_depth_frame_time = time.process_time()
            except:
                pass
            depthFrame = None
        except:
            pass
        depthFrameRef = None
        depthFrameEventData = None

  
    def handle_body_arrived(self, handle_index):
        bodyFrameEventData = self._body_frame_reader.GetFrameArrivedEventData(self._handles[handle_index])
        bofyFrameRef = bodyFrameEventData.FrameReference
        try:
            bodyFrame = bofyFrameRef.AcquireFrame()

            try: 
                with self._body_frame_lock:
                    bodyFrame.GetAndRefreshBodyData(self._body_frame_data_capacity, self._body_frame_data)
                    self._body_frame_bodies = KinectBodyFrameData(bodyFrame, self._body_frame_data, self.max_body_count)
                    self._last_body_frame_time = time.process_time()

                # need these 2 lines as a workaround for handling IBody referencing exception 
                self._body_frame_data = None
                self._body_frame_data = ctypes.cast(self._body_frame_data_type(), ctypes.POINTER(ctypes.POINTER(IBody)))

            except:
                pass
                            
            bodyFrame = None
        except:
            pass
        bofyFrameRef = None
        bodyFrameEventData = None


    def handle_body_index_arrived(self, handle_index):
        bodyIndexFrameEventData = self._body_index_frame_reader.GetFrameArrivedEventData(self._handles[handle_index])
        bodyIndexFrameRef = bodyIndexFrameEventData.FrameReference
        try:
            bodyIndexFrame = bodyIndexFrameRef.AcquireFrame()
            try:
                with self._body_index_frame_lock:
                    bodyIndexFrame.CopyFrameDataToArray(self._body_index_frame_data_capacity, self._body_index_frame_data)
                    self._last_body_index_frame_time = time.process_time()
            except: 
                pass
            bodyIndexFrame = None
        except:
            pass
        bodyIndexFrame = None
        bodyIndexFrameEventData = None

    def handle_infrared_arrived(self, handle_index):
        infraredFrameEventData = self._infrared_frame_reader.GetFrameArrivedEventData(self._handles[handle_index])
        infraredFrameRef = infraredFrameEventData.FrameReference
        try:
            infraredFrame = infraredFrameRef.AcquireFrame()
            try:
                with self._infrared_frame_lock:
                    infraredFrame.CopyFrameDataToArray(self._infrared_frame_data_capacity, self._infrared_frame_data)
                    self._last_infrared_frame_time = time.process_time()
            except:
                pass
            infraredFrame = None
        except:
            pass
        infraredFrameRef = None
        infraredFrameEventData = None

    def handle_long_exposure_infrared_arrived(self, handle_index):
        pass 

    def handle_audio_arrived(self, handle_index):
        pass 



class KinectBody(object): 
    def __init__(self, body = None):
        self.is_restricted = False
        self.tracking_id = -1

        self.is_tracked = False 
        
        if body is not None: 
            self.is_tracked = body.IsTracked

        if self.is_tracked:
            self.is_restricted = body.IsRestricted
            self.tracking_id = body.TrackingId
            self.engaged = body.Engaged
            self.lean = body.Lean
            self.lean_tracking_state = body.LeanTrackingState
            self.hand_left_state = body.HandLeftState
            self.hand_left_confidence = body.HandLeftConfidence
            self.hand_right_state = body.HandRightState
            self.hand_right_confidence = body.HandRightConfidence
            self.clipped_edges = body.ClippedEdges

            joints = ctypes.POINTER(PyKinectV2._Joint)
            joints_capacity = ctypes.c_uint(PyKinectV2.JointType_Count)
            joints_data_type = PyKinectV2._Joint * joints_capacity.value
            joints = ctypes.cast(joints_data_type(), ctypes.POINTER(PyKinectV2._Joint))
            body.GetJoints(PyKinectV2.JointType_Count, joints)
            self.joints = joints

            joint_orientations = ctypes.POINTER(PyKinectV2._JointOrientation)
            joint_orientations_data_type = PyKinectV2._JointOrientation * joints_capacity.value
            joint_orientations = ctypes.cast(joint_orientations_data_type(), ctypes.POINTER(PyKinectV2._JointOrientation))
            body.GetJointOrientations(PyKinectV2.JointType_Count, joint_orientations)
            self.joint_orientations = joint_orientations 


class KinectBodyFrameData(object): 
    def __init__(self, bodyFrame, body_frame_data, max_body_count):
        self.bodies = None
        self.floor_clip_plane = None
        if bodyFrame is not None:
            self.floor_clip_plane = bodyFrame.FloorClipPlane
            self.relative_time = bodyFrame.RelativeTime

            self.bodies = numpy.ndarray((max_body_count), dtype=numpy.object)
            for i in range(0, max_body_count):
               self.bodies[i] = KinectBody(body_frame_data[i])

    def copy(self):
        res = KinectBodyFrameData(None, None, 0)
        res.floor_clip_plane = self.floor_clip_plane
        res.relative_time = self.relative_time
        res.bodies = numpy.copy(self.bodies)
        return res


# colors for drawing different bodies 
SKELETON_COLORS = [pygame.color.THECOLORS["red"], 
                  pygame.color.THECOLORS["blue"], 
                  pygame.color.THECOLORS["green"], 
                  pygame.color.THECOLORS["orange"], 
                  pygame.color.THECOLORS["purple"], 
                  pygame.color.THECOLORS["yellow"], 
                  pygame.color.THECOLORS["violet"]]


class TimesUpDialog(gui.Dialog):
    def __init__(self,**params):
        title = gui.Label("massagebox")
        
        width = 400
        height = 50
        doc = gui.Document(width=width)
        
        space = title.style.font.size(" ")
        
        doc.block(align=0)
        for word in """TimeIsUp_RecordingFinished""".split(" "): 
            doc.add(gui.Label(word))
            doc.space(space)
        doc.br(space[1])     
        gui.Dialog.__init__(self,title,gui.ScrollArea(doc,width,height))


class SkeletonControl(gui.Table): 
    run_capture = False
    mydatetime = "0-0-0-0-0-0-0"
    status = "When there is a skeleton picture, click 'start' to record"
    textarea = gui.TextArea(value=status, width=500, height=20)
    save_info_textarea = gui.TextArea(value='', width=500, height=20)
    save_pic_textarea = gui.TextArea(value='', width=500, height=20)
    name = gui.Input(value='',size=16)
    kl_result = gui.Input(value='',size=16)
    save_pic = False
    sex = gui.Input(value='',size=16)
    age = gui.Input(value='',size=16)
    height = gui.Input(value='',size=16)
    weight = gui.Input(value='',size=16)
    

    def __init__(self,**params):
        gui.Table.__init__(self,**params)

        fg = (0,0,0)
        self.timesup_dialog = TimesUpDialog()
        self.tr()
        self.td(gui.Label("Skeleton GUI",color=fg),colspan=2)
        
        self.tr()
        self.td(gui.Label("Name: ",color=fg),align=1) 
        self.td(self.name,colspan=3)
        
        self.tr()
        self.td(gui.Label("KL_Result: ",color=fg),align=1)
        self.td(self.kl_result,colspan=3)

        self.tr()
        self.td(gui.Label("Sex: ",color=fg),align=1)
        self.td(self.sex,colspan=3)
    
        self.tr()
        self.td(gui.Label("Age: ",color=fg),align=1)
        self.td(self.age,colspan=3)
        
        self.tr()
        self.td(gui.Label("Height(m): ",color=fg),align=1)
        self.td(self.height,colspan=3)

        self.tr()
        self.td(gui.Label("Weight(kg): ",color=fg),align=1)
        self.td(self.weight,colspan=3)
        
        save_info_btn = gui.Button("Save Info")
        save_info_btn.connect(gui.CLICK, self.click_save_info_btn)
        self.tr()
        self.td(save_info_btn, colspan=1)
        
        start_btn = gui.Button("Start")
        start_btn.connect(gui.CLICK, self.click_start_btn)
        self.td(start_btn,colspan=1)

        stop_btn = gui.Button("Stop")
        stop_btn.connect(gui.CLICK, self.click_stop_btn)
        self.td(stop_btn,colspan=1)

        save_pic_btn = gui.Button("Save Pic")
        save_pic_btn.connect(gui.CLICK, self.click_save_pic_btn)
        self.td(save_pic_btn,colspan=1)

        self.tr()
        self.td(gui.Label("MainMassageBox: ",color=fg),align=1)
        self.td(SkeletonControl.textarea, colspan=4)

        self.tr()
        self.td(gui.Label("SaveInfoMassageBox: ",color=fg),align=1)
        self.td(SkeletonControl.save_info_textarea, colspan=4) 

        self.tr()
        self.td(gui.Label("SavePicMassageBox: ",color=fg),align=1)
        self.td(SkeletonControl.save_pic_textarea, colspan=4)
    

    def lan_broadcast_msg(self, msg):
        msg = msg.encode()
        host = "<broadcast>" # broadcast
        port = 6666
        addr = (host, port)
        UDPSock = socket(AF_INET, SOCK_DGRAM)
        UDPSock.bind(("", 0))
        UDPSock.setsockopt(SOL_SOCKET, SO_BROADCAST, 1)
        UDPSock.sendto(msg, addr)
        UDPSock.close()

    def click_save_info_btn(self):
        try:
            wb = load_workbook("patients_info.xlsx")
            sheet = wb.active
            nrows = sheet.max_row
            if self.name.value == sheet.cell(nrows, 1).value and self.kl_result.value == sheet.cell(nrows, 2).value and self.sex.value == sheet.cell(nrows, 3).value \
                and self.age.value == sheet.cell(nrows, 4).value and self.height.value == sheet.cell(nrows, 5).value and self.weight.value == sheet.cell(nrows, 6).value:
                self.save_info_textarea.value = '"' + self.name.value + '"' + " infomation already saved"
            elif self.name.value == "" and self.kl_result.value == "" and self.sex.value == "" \
                and self.age.value == "" and self.height.value == "" and self.weight.value == "":
                self.save_info_textarea.value = "please enter infomation first"
            else:
                sheet.cell(nrows+1, 1).value = self.name.value
                sheet.cell(nrows+1, 2).value = self.kl_result.value
                sheet.cell(nrows+1, 3).value = self.sex.value
                sheet.cell(nrows+1, 4).value = self.age.value
                sheet.cell(nrows+1, 5).value = self.height.value
                sheet.cell(nrows+1, 6).value = self.weight.value
                try:
                    wb.save(filename="patients_info.xlsx")
                    self.save_info_textarea.value = '"' + self.name.value + '"' + " infomation saved"
                except:
                    self.save_info_textarea.value = 'please close "patients_info.xlsx" first'

        except:
            wb = Workbook()
            wb = Workbook()
            sheet = wb.active
            sheet['A1'] = "姓名"
            sheet['B1'] = "膝关节炎KL分级诊断结果"
            sheet['C1'] = "性别"
            sheet['D1'] = "年龄"
            sheet['E1'] = "身高"
            sheet['F1'] = "体重"
            nrows = sheet.max_row
            if self.name.value == "" and self.kl_result.value == "" and self.sex.value == "" \
                and self.age.value == "" and self.height.value == "" and self.weight.value == "":
                self.save_info_textarea.value = "please enter infomation first"
            else:
                sheet.cell(nrows+1, 1).value = self.name.value
                sheet.cell(nrows+1, 2).value = self.kl_result.value
                sheet.cell(nrows+1, 3).value = self.sex.value
                sheet.cell(nrows+1, 4).value = self.age.value
                sheet.cell(nrows+1, 5).value = self.height.value
                sheet.cell(nrows+1, 6).value = self.weight.value
                try:
                    wb.save(filename="patients_info.xlsx")
                    self.save_info_textarea.value =  '"' + self.name.value + '"' + " infomation saved"
                except:
                    self.save_info_textarea.value = 'please close "patients_info.xlsx" first'


    def click_start_btn(self):
        if SkeletonControl.run_capture:
            self.textarea.value = "Recording Already Running"
            return
        else:
            msg = "startrecording" +  "-" + self.name.value + '-' + self.kl_result.value + '-' + self.sex.value + "-"  + \
                self.age.value + "-" + self.height.value + "-" + self.weight.value
            for i in range(10):
                self.lan_broadcast_msg(msg)
            
            self.s = Myscheduler(time.time, time.sleep)
            self.s.enter(10,1,self.times_up, ())
            t=threading.Thread(target=self.s.run)
            t.start()    

    def click_stop_btn(self):
        if SkeletonControl.run_capture:
            SkeletonControl.run_capture = False
            self.s.stop()
            self.textarea.value = "recording stopped" 
        else:
            self.textarea.value = "start recording first"
            # self.timesup_dialog.open() 

    def click_save_pic_btn(self):
        SkeletonControl.save_pic = True
               
    def get_datetime_string(self):
        string = str(datetime.now())
        mydate, mytime = string.split(" ")
        hour, minute, second = mytime.split(":")
        second = float(second)
        millisecond = str(int((second - int(second)) * 1000))
        second = str(int(second))
        return mydate + "-" + hour + '-' + minute + '-' +  second + '-' + millisecond

    def times_up(self):
        # self.run_capture = False
        msg = 'timesup'
        for i in range(10):
            self.lan_broadcast_msg(msg)
        self.textarea.value = "times up, recording stopped"
        self.timesup_dialog.open()

class Myscheduler(sched.scheduler):
    def __init__(self, timefunc=_time, delayfunc=time.sleep):
        sched.scheduler.__init__(self, timefunc=_time, delayfunc=time.sleep)
        self._running = True

    def stop(self):
        self._running = False

    def run(self, blocking=True):
        """Execute events until the queue is empty.
        If blocking is False executes the scheduled events due to
        expire soonest (if any) and then return the deadline of the
        next scheduled call in the scheduler.
        When there is a positive delay until the first event, the
        delay function is called and the event is left in the queue;
        otherwise, the event is removed from the queue and executed
        (its action function is called, passing it the argument).  If
        the delay function returns prematurely, it is simply
        restarted.
        It is legal for both the delay function and the action
        function to modify the queue or to raise an exception;
        exceptions are not caught but the scheduler's state remains
        well-defined so run() may be called again.
        A questionable hack is added to allow other threads to run:
        just after an event is executed, a delay of 0 is executed, to
        avoid monopolizing the CPU when other threads are also
        runnable.
        """
        # localize variable access to minimize overhead
        # and to improve thread safety
        lock = self._lock
        q = self._queue
        delayfunc = self.delayfunc
        timefunc = self.timefunc
        pop = heapq.heappop
        while self._running:
            with lock:
                if not q:
                    break
                time, priority, action, argument, kwargs = q[0]
                now = timefunc()
                if time > now:
                    delay = True
                else:
                    delay = False
                    pop(q)
            if delay:
                if not blocking:
                    return time - now
                delayfunc(time - now)
            else:
                action(*argument, **kwargs)
                delayfunc(0)   # Let other threads run


class SocketReceiver:
    def get_socket_data(self):
        host = ""
        port = 6666
        buf = 1024
        addr = (host, port)
        UDPSock = socket(AF_INET, SOCK_DGRAM)
        UDPSock.bind(addr)
        while True:
            (data, addr) = UDPSock.recvfrom(buf)
            data = data.decode()
            data = data.split('-')
            print(data)
            
            if data[0] == "startrecording":
                SkeletonControl.run_capture = True           
                SkeletonControl.mydatetime = self.get_datetime_string()
                SkeletonControl.name.value = data[1]
                SkeletonControl.kl_result.value = data[2]
                SkeletonControl.sex.value = data[3]
                SkeletonControl.age.value = data[4]
                SkeletonControl.height.value = data[5]
                SkeletonControl.weight.value = data[6]
            if data[0] == "timesup":
                SkeletonControl.run_capture = False
                SkeletonControl.textarea.value = "times up, recording stopped"
    
    def get_datetime_string(self):
        string = str(datetime.now())
        mydate, mytime = string.split(" ")
        hour, minute, second = mytime.split(":")
        second = float(second)
        millisecond = str(int((second - int(second)) * 1000))
        second = str(int(second))
        return mydate + "-" + hour + '-' + minute + '-' +  second + '-' + millisecond
    

    def run(self):
        self.get_socket_data()
        


class BodyGameRuntime(object):
    
    def __init__(self):
        pygame.init()

        # Used to manage how fast the screen updates
        self._clock = pygame.time.Clock()

        # Set the width and height of the screen [width, height]
        self._infoObject = pygame.display.Info()
        self._screen = pygame.display.set_mode(((self._infoObject.current_w >> 1) + 300, (self._infoObject.current_h >> 1) + 300), 
                                               pygame.HWSURFACE|pygame.DOUBLEBUF|pygame.RESIZABLE, 32)
        # pygame.display.set_caption("Kinect-based Gait Data Acquisition Software")
        pygame.display.set_caption("基于Kinect的步态数据采集软件v1.0")

        # Loop until the user clicks the close button.
        self._done = False

        # Used to manage how fast the screen updates
        self._clock = pygame.time.Clock()

        # Kinect runtime object, we want only color and body frames 
        self._kinect = PyKinectRuntime(PyKinectV2.FrameSourceTypes_Color | PyKinectV2.FrameSourceTypes_Body)

        # back buffer surface for getting Kinect color frames, 32bit color, width and height equal to the Kinect color frame size
        self._frame_surface = pygame.Surface((self._kinect.color_frame_desc.Width, self._kinect.color_frame_desc.Height), 0, 32)
        
        # surface to draw skeleton
        self._skeleton_surface = pygame.Surface((self._kinect.color_frame_desc.Width, self._kinect.color_frame_desc.Height), 0, 32)

        self._gui_surface = pygame.Surface((self._screen.get_width(), self._screen.get_height() // 2), 0, 32)
        

        # here we will store skeleton data 
        self._bodies = None

        # self.app = gui.App()
        
        self.app = gui.Desktop()
        # self.app.screen = self._gui_surface



        self.skeletonCtrl =SkeletonControl()

        # self.c = gui.Container(align=-1,valign=-1, x=self._screen.get_width() // 2, y=self._frame_surface.get_height() // 2)
        self.c = gui.Container(align=-1,valign=-1, x=0, y=0)

        
        # self.c.add(self.skeletonCtrl,0, 0)
        self.c.add(self.skeletonCtrl,300, 0)
        
        # self.app.init(widget=self.c)#, screen=self._gui_surface)  # area=(0, self._screen.get_height() // 2, self._screen.get_width(), self._screen.get_height() // 2))
        # self.app.init(widget=self.c, screen=self._screen, area=pygame.Rect(0, self._screen.get_height() // 2, self._screen.get_width(), self._screen.get_height() // 2))
        self.h_to_w = float(self._frame_surface.get_height()) / self._frame_surface.get_width()
            # print self._frame_surface.get_height(), self._frame_surface.get_width()
        self.target_height = int(self.h_to_w * self._screen.get_width() // 2)
        self.app.init(widget=self.c, screen=self._screen, area=pygame.Rect(0, self.target_height, self._screen.get_width(), self._screen.get_height() - self.target_height))



    def draw_body_bone(self, joints, jointPoints, color, joint0, joint1):
        joint0State = joints[joint0].TrackingState
        joint1State = joints[joint1].TrackingState

        # both joints are not tracked
        if (joint0State == PyKinectV2.TrackingState_NotTracked) or (joint1State == PyKinectV2.TrackingState_NotTracked): 
            return

        # both joints are not *really* tracked
        if (joint0State == PyKinectV2.TrackingState_Inferred) and (joint1State == PyKinectV2.TrackingState_Inferred):
            return

        # ok, at least one is good 
        start = (jointPoints[joint0].x, jointPoints[joint0].y)
        end = (jointPoints[joint1].x, jointPoints[joint1].y)

        try:
            pygame.draw.line(self._skeleton_surface, color, start, end, 8)
        except: # need to catch it due to possible invalid positions (with inf)
            pass

    def draw_body(self, joints, jointPoints, color):
        # Torso
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_Head, PyKinectV2.JointType_Neck)
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_Neck, PyKinectV2.JointType_SpineShoulder)
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_SpineShoulder, PyKinectV2.JointType_SpineMid)
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_SpineMid, PyKinectV2.JointType_SpineBase)
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_SpineShoulder, PyKinectV2.JointType_ShoulderRight)
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_SpineShoulder, PyKinectV2.JointType_ShoulderLeft)
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_SpineBase, PyKinectV2.JointType_HipRight)
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_SpineBase, PyKinectV2.JointType_HipLeft)
    
        # Right Arm    
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_ShoulderRight, PyKinectV2.JointType_ElbowRight)
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_ElbowRight, PyKinectV2.JointType_WristRight)
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_WristRight, PyKinectV2.JointType_HandRight)
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_HandRight, PyKinectV2.JointType_HandTipRight)
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_WristRight, PyKinectV2.JointType_ThumbRight)

        # Left Arm
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_ShoulderLeft, PyKinectV2.JointType_ElbowLeft)
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_ElbowLeft, PyKinectV2.JointType_WristLeft)
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_WristLeft, PyKinectV2.JointType_HandLeft)
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_HandLeft, PyKinectV2.JointType_HandTipLeft)
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_WristLeft, PyKinectV2.JointType_ThumbLeft)

        # Right Leg
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_HipRight, PyKinectV2.JointType_KneeRight)
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_KneeRight, PyKinectV2.JointType_AnkleRight)
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_AnkleRight, PyKinectV2.JointType_FootRight)

        # Left Leg
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_HipLeft, PyKinectV2.JointType_KneeLeft)
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_KneeLeft, PyKinectV2.JointType_AnkleLeft)
        self.draw_body_bone(joints, jointPoints, color, PyKinectV2.JointType_AnkleLeft, PyKinectV2.JointType_FootLeft)


    def draw_color_frame(self, frame, target_surface):
        target_surface.lock()
        address = self._kinect.surface_as_array(target_surface.get_buffer())
        ctypes.memmove(address, frame.ctypes.data, frame.size)
        del address
        target_surface.unlock()

    

    def run(self):
        # run a new thread to receive socket data 
        socket_receiver = SocketReceiver()
        receiver = threading.Thread(target=socket_receiver.run)
        receiver.start()
        
        # -------- Main Program Loop -----------
        while not self._done:
            # self.app.init(widget=self.c, screen=self._screen, area=pygame.Rect(0, self._screen.get_height() // 2, self._screen.get_width(), self._screen.get_height() // 2))

            # --- Main event loop
            for event in pygame.event.get(): # User did something
                if event.type == pygame.QUIT: # If user clicked close
                    self._done = True # Flag that we are done so we exit this loop

                elif event.type == pygame.VIDEORESIZE: # window resized
                    self._screen = pygame.display.set_mode(event.dict['size'], 
                                               pygame.HWSURFACE|pygame.DOUBLEBUF|pygame.RESIZABLE, 32)
                    self.target_height = int(self.h_to_w * self._screen.get_width() // 2)
                    # self.app.init(widget=self.c, screen=self._screen, area=pygame.Rect(0, self._screen.get_height() // 2 + 50, self._screen.get_width(), self._screen.get_height() // 2 - 50))
                    self.app.init(widget=self.c, screen=self._screen, area=pygame.Rect(0, self.target_height, self._screen.get_width(), self._screen.get_height() - self.target_height))

                    
                else:
                    self.app.event(event)
            # print self._frame_surface.get_rect()      
            # print self._skeleton_surface.get_rect()      

            # --- Game logic should go here

            # --- Getting frames and drawing  
            # --- Woohoo! We've got a color frame! Let's fill out back buffer surface with frame's data 
            if self._kinect.has_new_color_frame():
                frame = self._kinect.get_last_color_frame()
                self.draw_color_frame(frame, self._frame_surface)
                frame = None


            # --- Cool! We have a body frame, so can get skeletons
            if self._kinect.has_new_body_frame(): 
                self._bodies = self._kinect.get_last_body_frame()
            # else:
            #     self._bodies = None

            # --- draw skeletons to _frame_surface and write skeleton data to txt
            if self._bodies is not None: 
                counter = 0
                for i in range(0, self._kinect.max_body_count):
                    body = self._bodies.bodies[i]
                    if not body.is_tracked: 
                        counter += 1
                        if counter == self._kinect.max_body_count:
                            SkeletonControl.textarea.value = "no skeleton data to record"
                        continue
                    if SkeletonControl.textarea.value == "no skeleton data to record":
                        SkeletonControl.textarea.value = "click 'start' to record"
                    joints = body.joints
                    # convert joint coordinates to color space 
                    joint_points = self._kinect.body_joints_to_color_space(joints)
                    self.draw_body(joints, joint_points, SKELETON_COLORS[i])

                    # save skeleton data to .txt file 
                    if SkeletonControl.run_capture == True:
                        SkeletonControl.textarea.value = "recording......"
                        with open(SkeletonControl.mydatetime + "-" + SkeletonControl.name.value + "-" + SkeletonControl.kl_result.value + ".txt",'a') as f:
                            for i in range(PyKinectV2.JointType_Count):                              
                                f.write('{:.7f}'.format(joints[i].Position.x) + ' ' + '{:.7f}'.format(joints[i].Position.y) + ' ' + '{:.7f}'.format(joints[i].Position.z) + ' ')                          
                            f.write("\n")

            # --- copy back buffer surface pixels to the screen, resize it if needed and keep aspect ratio
            # --- (screen size may be different from Kinect's color frame size) 
            # h_to_w = float(self._frame_surface.get_height()) / self._frame_surface.get_width()
            # print self._frame_surface.get_height(), self._frame_surface.get_width()
            # target_height = int(self.h_to_w * self._screen.get_width() // 2)
            surface_to_draw = pygame.transform.scale(self._frame_surface, (self._screen.get_width() // 2, self.target_height))
            if SkeletonControl.save_pic == True:
                pygame.image.save(self._frame_surface, self.skeletonCtrl.get_datetime_string() + "-" +  SkeletonControl.name.value + '-' + SkeletonControl.kl_result.value + '.png')
                SkeletonControl.save_pic_textarea.value = '"' + self.skeletonCtrl.get_datetime_string() + "-" + SkeletonControl.name.value + '-' + SkeletonControl.kl_result.value + '"' + " picture saved"
                SkeletonControl.save_pic = False
            skeleton_surface_to_draw = pygame.transform.scale(self._skeleton_surface, (self._screen.get_width() // 2, self.target_height))
            # gui_surface_to_draw = pygame.transform.scale(self._gui_surface, (self._screen.get_width(), self._screen.get_height() // 2))
            self._screen.blit(surface_to_draw, (0,0))
            self._screen.blit(skeleton_surface_to_draw, (self._screen.get_width() // 2, 0))
            # self._screen.blit(gui_surface_to_draw, (0, self._screen.get_height() // 2))
            surface_to_draw = None
            skeleton_surface_to_draw = None
            # gui_surface_to_draw = None
            self._skeleton_surface.fill((0, 0, 0))
            
            pygame.display.update()
            self.app.paint()

            # --- Go ahead and update the screen with what we've drawn.
            pygame.display.flip()

            # --- Limit to 30 frames per second
            self._clock.tick(30)

        # Close our Kinect sensor, close the window and quit.
        self._kinect.close()
        pygame.quit()


__main__ = "Kinect v2 Body Game"
game = BodyGameRuntime()
game.run()
